"""
KGL 출고 오더 자동화 처리 스크립트
- SungwooToHT: 오더 수량 원천 파일
- FromSungwoo: 정렬 기준 템플릿
- KGL 출고 오더 포맷: 최종 출력 파일
"""

import io
import pandas as pd
from openpyxl import load_workbook


# ─── 파일 자동 분류 ───────────────────────────────────────────────────────────

FILE_SIGNATURES = {
    "sungwoo_to_ht": {"No.", "Quantity", "Each Price"},
    "from_sungwoo":  {"Item No.", "Available Qty.", "Each Price 1"},
    "kgl_format":    {"SAP Code", "입출고수량"},
}

# KGL 헤더의 한글 컬럼 인식 보조 (인코딩 이슈 대비 영문 fallback)
KGL_FALLBACK_COLS = {"SAP Code", "Item Code", "Stock"}


def classify_file(path: str) -> str:
    """파일 헤더를 스캔하여 유형(sungwoo_to_ht / from_sungwoo / kgl_format) 반환."""
    filename = str(path).lower()
    if "sungwooToHT".lower() in filename or "sungwoo_to_ht" in filename:
        return "sungwoo_to_ht"
    if "fromsungwoo" in filename.replace(" ", "").replace("_", ""):
        return "from_sungwoo"
    if "kgl" in filename or "inventory" in filename:
        return "kgl_format"

    # 파일명 단서가 없으면 컬럼으로 판별
    try:
        header_row = _find_header_row(path)
        raw = pd.read_excel(path, header=header_row, nrows=0)
        cols = set(str(c).strip() for c in raw.columns)
        for ftype, sig in FILE_SIGNATURES.items():
            if sig.issubset(cols):
                return ftype
        if KGL_FALLBACK_COLS.issubset(cols):
            return "kgl_format"
    except Exception:
        pass

    raise ValueError(f"파일 유형을 자동으로 분류할 수 없습니다: {path}")


# ─── 동적 헤더 행 탐색 ────────────────────────────────────────────────────────

_HEADER_HINTS = {
    "sungwoo_to_ht": {"No.", "Quantity"},
    "from_sungwoo":  {"Item No.", "Available Qty."},
    "kgl_format":    {"SAP Code"},
}


def _find_header_row(path: str, file_type: str | None = None, max_scan: int = 20) -> int:
    """핵심 키 컬럼이 처음 등장하는 행 인덱스를 반환한다 (0-based)."""
    hints = set()
    if file_type and file_type in _HEADER_HINTS:
        hints = _HEADER_HINTS[file_type]

    raw = pd.read_excel(path, header=None, nrows=max_scan)
    for i, row in raw.iterrows():
        cells = {str(v).strip() for v in row if pd.notna(v)}
        if hints and hints.issubset(cells):
            return i
        # hints 없을 때: 첫 번째 유형의 시그니처 중 하나라도 매칭
        for sig in FILE_SIGNATURES.values():
            if sig.issubset(cells):
                return i
    return 0


# ─── 코드 정규화 ──────────────────────────────────────────────────────────────

def normalize_code(code) -> str:
    """상품 코드를 'H' 접두사 없는 순수 숫자 문자열로 정규화한다."""
    s = str(code).strip().upper()
    if s.startswith("H"):
        s = s[1:]
    # 소수점 제거 (111069.0 → 111069)
    if "." in s:
        try:
            s = str(int(float(s)))
        except ValueError:
            pass
    return s


# ─── Step 1~2: SungwooToHT → 수량 딕셔너리 ──────────────────────────────────

def load_order_qty(path: str) -> dict[str, float]:
    """SungwooToHT 파일에서 {정규화코드: 합산수량} 딕셔너리를 반환한다."""
    hr = _find_header_row(path, "sungwoo_to_ht")
    df = pd.read_excel(path, header=hr, dtype={"No.": str})
    df = df[df["No."].notna()].copy()
    df["_key"] = df["No."].apply(normalize_code)
    df["Quantity"] = pd.to_numeric(df["Quantity"], errors="coerce").fillna(0)
    return df.groupby("_key")["Quantity"].sum().to_dict()


# ─── Step 3: FromSungwoo → 정렬 기준 키 리스트 ───────────────────────────────

def load_sort_order(path: str) -> list[str]:
    """FromSungwoo 파일의 Item No. 순서를 정규화 코드 리스트로 반환한다."""
    hr = _find_header_row(path, "from_sungwoo")
    df = pd.read_excel(path, header=hr, dtype={"Item No.": str})
    df = df[df["Item No."].notna()].copy()
    return [normalize_code(v) for v in df["Item No."]]


# ─── Step 4~5: KGL 포맷 파일에 수량 기록 ────────────────────────────────────

def _find_kgl_header_row_in_wb(ws) -> int:
    """openpyxl 워크시트에서 'SAP Code' 가 있는 행 번호(1-based)를 반환한다."""
    for row in ws.iter_rows():
        for cell in row:
            if str(cell.value).strip() == "SAP Code":
                return cell.row
    raise ValueError("KGL 파일에서 'SAP Code' 헤더를 찾을 수 없습니다.")


def write_qty_to_kgl(
    kgl_path: str,
    order_qty: dict[str, float],
    sort_order: list[str],
    output_path: str | None = None,
) -> bytes:
    """
    KGL 출고 오더 포맷 파일의 '입출고수량' 열에 수량을 기록하고
    바이너리 스트림을 반환한다. 기존 서식은 절대 훼손하지 않는다.
    """
    wb = load_workbook(kgl_path)
    ws = wb.active

    header_row = _find_kgl_header_row_in_wb(ws)

    # 헤더 행에서 SAP Code · 입출고수량 열 번호 탐색
    sap_col = qty_col = None
    for cell in ws[header_row]:
        val = str(cell.value).strip() if cell.value is not None else ""
        if val == "SAP Code":
            sap_col = cell.column
        if val == "입출고수량":
            qty_col = cell.column

    if sap_col is None:
        raise ValueError("KGL 파일에서 'SAP Code' 열을 찾을 수 없습니다.")
    if qty_col is None:
        raise ValueError("KGL 파일에서 '입출고수량' 열을 찾을 수 없습니다.")

    # FromSungwoo 정렬 순서 기반 매핑 테이블 생성
    # sort_order 에 없는 코드도 처리할 수 있도록 전체 order_qty 사용
    sorted_qty: dict[str, float] = {}
    for code in sort_order:
        if code in order_qty:
            sorted_qty[code] = order_qty[code]
    # sort_order 에 없지만 오더에 있는 항목도 누락 없이 포함
    for code, qty in order_qty.items():
        if code not in sorted_qty:
            sorted_qty[code] = qty

    # 데이터 행 순회하며 수량 기록 (동일 SAP Code 첫 번째 행에만 기록)
    written: set[str] = set()
    for row_idx in range(header_row + 1, ws.max_row + 1):
        sap_cell = ws.cell(row=row_idx, column=sap_col)
        if sap_cell.value is None:
            continue
        norm = normalize_code(sap_cell.value)
        if norm in sorted_qty and norm not in written:
            qty_cell = ws.cell(row=row_idx, column=qty_col)
            qty_cell.value = sorted_qty[norm]
            written.add(norm)

    buf = io.BytesIO()
    wb.save(buf)
    if output_path:
        with open(output_path, "wb") as f:
            f.write(buf.getvalue())
    buf.seek(0)
    return buf.read()


# ─── 메인 진입점 ─────────────────────────────────────────────────────────────

def process_order_files(
    file_paths: list[str],
    output_path: str | None = None,
) -> bytes:
    """
    3개의 파일 경로를 순서 무관하게 받아 자동 분류 후 처리하고
    완성된 KGL 파일을 바이너리로 반환한다.

    Parameters
    ----------
    file_paths   : [path1, path2, path3] — 순서 무관
    output_path  : 결과를 파일로도 저장할 경우 경로 지정 (선택)

    Returns
    -------
    bytes : 완성된 엑셀 파일 바이너리
    """
    classified: dict[str, str] = {}
    for p in file_paths:
        ftype = classify_file(p)
        if ftype in classified:
            raise ValueError(f"동일 유형({ftype})의 파일이 2개 이상 업로드 되었습니다.")
        classified[ftype] = p

    missing = {"sungwoo_to_ht", "from_sungwoo", "kgl_format"} - set(classified)
    if missing:
        raise ValueError(f"다음 유형의 파일이 누락되었습니다: {missing}")

    print("[1/3] 오더 수량 취합 중... (SungwooToHT)")
    order_qty = load_order_qty(classified["sungwoo_to_ht"])
    print(f"      → {len(order_qty)}개 품목 수량 로드 완료")

    print("[2/3] 정렬 기준 로드 중... (FromSungwoo)")
    sort_order = load_sort_order(classified["from_sungwoo"])
    print(f"      → {len(sort_order)}개 품목 순서 기준 로드 완료")

    print("[3/3] KGL 포맷에 수량 기록 중...")
    result = write_qty_to_kgl(
        classified["kgl_format"], order_qty, sort_order, output_path
    )
    matched = sum(1 for c in sort_order if normalize_code(c) in order_qty)
    print(f"      → 매칭 성공: {matched}개 / 오더 품목: {len(order_qty)}개")
    if output_path:
        print(f"      → 저장 완료: {output_path}")
    return result


# ─── 직접 실행 예시 ───────────────────────────────────────────────────────────

if __name__ == "__main__":
    import sys
    import os

    # 기본 경로 (다운로드 폴더 기준)
    base = r"C:\Users\rlehd\Downloads"
    files = [
        os.path.join(base, "SungwooToHT.xlsx"),
        os.path.join(base, "FromSungwoo.xlsx"),
        os.path.join(base, "KGL 출고 오더 포맷.xlsx"),
    ]
    out = os.path.join(base, "KGL_출고오더_완성.xlsx")

    try:
        process_order_files(files, output_path=out)
        print(f"\n완료! 결과 파일: {out}")
    except Exception as e:
        print(f"\n오류 발생: {e}", file=sys.stderr)
        sys.exit(1)
